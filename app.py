import os
import uuid
from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from io import BytesIO

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max


@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Page non trouvée'}), 404


@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': 'Fichier trop volumineux (max 16MB)'}), 413


@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': 'Erreur interne du serveur'}), 500


@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({'error': str(e)}), 500

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)


def safe_path(file_id):
    """Validate file_id is a UUID and return safe file path."""
    if not file_id:
        return None
    try:
        uuid.UUID(file_id)
    except (ValueError, AttributeError):
        return None
    return os.path.join(UPLOAD_DIR, f'{file_id}.xlsx')


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext != '.xlsx':
        return jsonify({'error': 'Format non supporté. Veuillez utiliser un fichier .xlsx'}), 400

    file_id = str(uuid.uuid4())
    filepath = safe_path(file_id)
    file.save(filepath)

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active

        headers = []
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for i, val in enumerate(first_row):
            headers.append(str(val) if val is not None else f'Colonne {i + 1}')

        preview = []
        for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
            padded = []
            for i in range(len(headers)):
                v = row[i] if i < len(row) else None
                padded.append(str(v) if v is not None else '')
            preview.append(padded)

        row_count = ws.max_row - 1
        wb.close()

        return jsonify({
            'file_id': file_id,
            'headers': headers,
            'row_count': max(row_count, 0),
            'preview': preview,
            'filename': file.filename
        })
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'error': f'Erreur lors de la lecture: {str(e)}'}), 400


@app.route('/api/process', methods=['POST'])
def process():
    data = request.get_json()
    file_id = data.get('file_id')
    col_index = data.get('column_index', 0)

    filepath = safe_path(file_id)
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'Fichier non trouvé. Veuillez réimporter.'}), 404

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    results = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        cell_value = row[col_index] if col_index < len(row) else None
        full_name = str(cell_value).strip() if cell_value is not None else ''

        if not full_name:
            results.append({
                'index': row_idx,
                'original': '',
                'parts': [],
                'splitAt': 0
            })
            continue

        parts = full_name.split()

        if len(parts) <= 1:
            split_at = len(parts)
        else:
            # Default heuristic: last word is Nom, rest is Prénom
            split_at = len(parts) - 1

        results.append({
            'index': row_idx,
            'original': full_name,
            'parts': parts,
            'splitAt': split_at
        })

    wb.close()
    return jsonify({'names': results})


@app.route('/api/download', methods=['POST'])
def download():
    data = request.get_json()
    file_id = data.get('file_id')
    col_index = data.get('column_index', 0)
    names = data.get('names', [])

    filepath = safe_path(file_id)
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'Fichier non trouvé. Veuillez réimporter.'}), 404

    original_wb = openpyxl.load_workbook(filepath)
    original_ws = original_wb.active

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = original_ws.title

    # Original headers
    orig_headers = []
    for cell in original_ws[1]:
        orig_headers.append(str(cell.value) if cell.value is not None else f'Colonne {cell.column}')

    # Build new headers: replace name column with Prénom + Nom
    new_headers = []
    for i, h in enumerate(orig_headers):
        if i == col_index:
            new_headers.append('Prénom')
            new_headers.append('Nom')
        else:
            new_headers.append(h)

    # Write headers
    ws.append(new_headers)

    # Build name lookup from client data
    name_lookup = {}
    for n in names:
        parts = n.get('parts', [])
        split_at = n.get('splitAt', 0)
        prenom = ' '.join(parts[:split_at])
        nom = ' '.join(parts[split_at:])
        name_lookup[n['index']] = (prenom, nom)

    # Write data rows
    for row_idx, row in enumerate(original_ws.iter_rows(min_row=2, values_only=True)):
        out_row = []
        for i, value in enumerate(row):
            if i == col_index:
                if row_idx in name_lookup:
                    prenom, nom = name_lookup[row_idx]
                else:
                    prenom, nom = '', ''
                out_row.append(prenom)
                out_row.append(nom)
            else:
                out_row.append(value)
        ws.append(out_row)

    original_wb.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='resultat_prenoms_noms.xlsx'
    )


@app.route('/api/cleanup', methods=['POST'])
def cleanup():
    data = request.get_json()
    file_id = data.get('file_id')
    filepath = safe_path(file_id)
    if filepath and os.path.exists(filepath):
        os.remove(filepath)
    return jsonify({'ok': True})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(host='0.0.0.0', port=port, debug=debug)
